#!/usr/bin/env python3
"""Quick test: qwen-flash generates Excel code, executes locally, checks output."""
import json, os, sys, subprocess, tempfile, shutil, re
sys.stdout.reconfigure(encoding='utf-8')
from openai import OpenAI

client = OpenAI(api_key=os.environ['TARGET_OPENAI_COMPATIBLE_API_KEY'],
                base_url=os.environ['TARGET_OPENAI_COMPATIBLE_BASE_URL'])

with open('data/spreadsheetbench_verified_400/dataset.json', encoding='utf-8') as f:
    items = json.load(f)

results = []
for item in items[:5]:
    ss_dir = f'data/spreadsheetbench_verified_400/{item["spreadsheet_path"]}'
    init_file = [f for f in os.listdir(ss_dir) if f.endswith('_init.xlsx')][0]

    prompt = f'''Write Python code using openpyxl to solve this task:

{item["instruction"]}

Input: input.xlsx
Output: output.xlsx

Requirements:
- INPUT_PATH = 'input.xlsx' at the top
- OUTPUT_PATH = 'output.xlsx' at the top
- Handle headers (first row may be column names)
- wb.save(OUTPUT_PATH) at the end

Output ONLY raw Python code. No markdown fences (no ```python). No explanation.'''

    try:
        resp = client.chat.completions.create(
            model='qwen-flash',
            messages=[{'role':'system','content':'You are a Python + openpyxl expert. Write ONLY Python code. No intro, no markdown.'},
                      {'role':'user','content': prompt}],
            max_tokens=2000,
            temperature=0)
        raw_code = resp.choices[0].message.content or ''
    except Exception as e:
        print(f'[{item["id"]}] API ERROR: {e}')
        continue

    # Clean code
    code = raw_code.strip()
    if code.startswith('```'):
        code = re.sub(r'^```\w*\n', '', code)
        code = re.sub(r'\n```\s*$', '', code)
        code = code.strip()

    # Execute
    workdir = tempfile.mkdtemp(prefix='ssb_')
    ok = False
    error = ''
    try:
        shutil.copy(f'{ss_dir}/{init_file}', f'{workdir}/input.xlsx')
        with open(f'{workdir}/solution.py', 'w', encoding='utf-8', newline='\n') as f:
            f.write(code)

        result = subprocess.run(
            [sys.executable, 'solution.py'],
            capture_output=True, text=True, encoding='utf-8',
            timeout=120, cwd=workdir
        )

        out_file = f'{workdir}/output.xlsx'
        if result.returncode == 0 and os.path.exists(out_file) and os.path.getsize(out_file) > 100:
            ok = True
            import openpyxl
            wb = openpyxl.load_workbook(out_file)
            for sn in wb.sheetnames[:1]:
                ws = wb[sn]
                print(f'  [{item["id"]}] ✅ rows={ws.max_row} cols={ws.max_column}')
        else:
            error = result.stderr[:200] if result.stderr else f'exit={result.returncode} no_output={not os.path.exists(out_file)}'
            print(f'  [{item["id"]}] ❌ {error}')
    except Exception as e:
        print(f'  [{item["id"]}] 💥 {e}')
    finally:
        shutil.rmtree(workdir, ignore_errors=True)

    results.append((item["id"], ok))

print(f'\nSuccess: {sum(1 for _,ok in results if ok)}/{len(results)}')
