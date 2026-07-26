#!/usr/bin/env python3
"""Simplified SpreadsheetBench rollout — qwen-flash + local execution.
Tests whether SkillOpt can optimize spreadsheet-coding skill.
"""
from __future__ import annotations
import json, os, sys, re, subprocess, tempfile, shutil, time, traceback
from concurrent.futures import ThreadPoolExecutor
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
sys.stdout.reconfigure(encoding='utf-8')

import openpyxl
import numpy as np
from openai import OpenAI
from skillopt.envs.spreadsheetbench.evaluator import _compare_cell_value, _transform_value

# ── Config ──────────────────────────────────────────────────────────────
client = OpenAI(api_key=os.environ.get('TARGET_OPENAI_COMPATIBLE_API_KEY', ''),
                base_url=os.environ.get('TARGET_OPENAI_COMPATIBLE_BASE_URL', ''))

DATA_ROOT = 'data/spreadsheetbench_verified_400'
CODE_TIMEOUT = 120  # seconds


# ── Execute generated code ──────────────────────────────────────────────

def execute_code(code: str, input_path: str, work_dir: str) -> str:
    """Execute Python code in work_dir. Returns 'ok' or error message."""
    # Clean code
    code = code.strip()
    code = re.sub(r'^```\w*\n', '', code)
    code = re.sub(r'\n```\s*$', '', code)

    script_path = os.path.join(work_dir, 'solution.py')
    with open(script_path, 'w', encoding='utf-8', newline='\n') as f:
        f.write(code)

    shutil.copy(input_path, os.path.join(work_dir, 'input.xlsx'))

    try:
        result = subprocess.run(
            [sys.executable, 'solution.py'],
            capture_output=True, text=True, encoding='utf-8',
            timeout=CODE_TIMEOUT, cwd=work_dir
        )
        out_file = os.path.join(work_dir, 'output.xlsx')
        if result.returncode == 0 and os.path.exists(out_file) and os.path.getsize(out_file) > 100:
            return 'ok'
        err = result.stderr[:300] or f'exit={result.returncode}'
        return err.strip()
    except subprocess.TimeoutExpired:
        return f'timeout after {CODE_TIMEOUT}s'
    except Exception as e:
        return f'{type(e).__name__}: {e}'


# ── Evaluate vs gold ────────────────────────────────────────────────────

def compare_with_gold(pred_path: str, gold_path: str) -> tuple[int, int]:
    """Return (cells_correct, cells_compared)."""
    try:
        wb_pred = openpyxl.load_workbook(pred_path, data_only=True)
        wb_gold = openpyxl.load_workbook(gold_path, data_only=True)
    except Exception:
        return 0, 1

    correct, total = 0, 0
    for sn in wb_gold.sheetnames:
        if sn not in wb_pred.sheetnames:
            continue
        ws_pred = wb_pred[sn]
        ws_gold = wb_gold[sn]
        for row_gold in ws_gold.iter_rows():
            for cell_gold in row_gold:
                cell_pred = ws_pred.cell(row=cell_gold.row, column=cell_gold.column)
                total += 1
                if _compare_cell_value(cell_pred.value, cell_gold.value):
                    correct += 1
    return correct, total if total > 0 else 1


# ── Prompt construction ─────────────────────────────────────────────────

SYSTEM_PROMPT = """You are a Python + openpyxl expert. Write concise, correct code.
Rules:
- INPUT_PATH = 'input.xlsx' at the top
- OUTPUT_PATH = 'output.xlsx' at the top
- Handle header rows (usually row 1)
- Use wb.save(OUTPUT_PATH) at the end
- No markdown fences, no explanation — ONLY the Python code"""


def build_prompt(instruction: str, skill: str) -> str:
    skill_block = f"{skill}\n\n" if skill.strip() else ""
    return f"""{skill_block}## Task
{instruction}

Write a Python script using openpyxl to complete this task.
Input file: input.xlsx
Output file: output.xlsx

## Code Requirements
{skill_block}
- Define INPUT_PATH = 'input.xlsx'
- Define OUTPUT_PATH = 'output.xlsx'
- Read the input Excel, process according to the task, save to output.
- wb.save(OUTPUT_PATH) at the very end.

Output ONLY the Python code."""


# ── Single-item rollout ─────────────────────────────────────────────────

def rollout_one(item: dict, skill: str) -> dict:
    """Run one SpreadsheetBench item. Returns {hard, soft, ...}."""
    ss_dir = os.path.join(DATA_ROOT, item['spreadsheet_path'])
    init_files = [f for f in os.listdir(ss_dir) if f.endswith('_init.xlsx')]
    gold_files = [f for f in os.listdir(ss_dir) if f.endswith('_golden.xlsx')]

    if not init_files:
        return {'hard': 0, 'soft': 0.0, 'fail_reason': 'no_init_file'}

    user_prompt = build_prompt(item['instruction'], skill)
    init_path = os.path.join(ss_dir, init_files[0])

    # Generate code
    try:
        resp = client.chat.completions.create(
            model='qwen-flash',
            messages=[{'role': 'system', 'content': SYSTEM_PROMPT},
                      {'role': 'user', 'content': user_prompt}],
            max_tokens=2000, temperature=0)
        code = resp.choices[0].message.content or ''
    except Exception as e:
        return {'hard': 0, 'soft': 0.0, 'fail_reason': f'api: {e}'}

    # Execute
    work_dir = tempfile.mkdtemp(prefix='ssb_')
    try:
        gold_path = os.path.join(ss_dir, gold_files[0]) if gold_files else None

        exec_result = execute_code(code, init_path, work_dir)
        if exec_result != 'ok':
            # Check if output was produced despite error
            pred_path = os.path.join(work_dir, 'output.xlsx')
            if gold_path and os.path.exists(pred_path) and os.path.getsize(pred_path) > 100:
                correct, total = compare_with_gold(pred_path, gold_path)
                acc = correct / total if total > 0 else 0.0
                if acc >= 0.90:
                    hard = 0.8; reason = f'ran_with_errors_but_acc={acc:.3f}'
                elif acc >= 0.50:
                    hard = 0.3; reason = f'ran_with_errors_acc={acc:.3f}'
                else:
                    hard = 0.05; reason = f'ran_with_errors_low_acc={acc:.3f}'
                return {'hard': hard, 'soft': acc, 'fail_reason': reason}
            return {'hard': 0, 'soft': 0.0, 'fail_reason': f'exec: {exec_result[:120]}'}

        # Compare with gold
        if gold_path:
            pred_path = os.path.join(work_dir, 'output.xlsx')
            correct, total = compare_with_gold(pred_path, gold_path)
            acc = correct / total if total > 0 else 0.0
            # Tiered scoring: hard = 1.0 for perfect, 0.5 for >90%, 0.25 for >70%
            if acc >= 0.99:
                hard = 1.0; reason = 'perfect'
            elif acc >= 0.90:
                hard = 0.5; reason = f'minor_diff'
            elif acc >= 0.70:
                hard = 0.25; reason = f'significant_diff'
            elif acc >= 0.40:
                hard = 0.1; reason = f'major_diff'
            else:
                hard = 0; reason = f'wrong'
            return {'hard': hard, 'soft': acc,
                    'fail_reason': reason + f' acc={acc:.3f} ({correct}/{total})'}
        else:
            return {'hard': 1, 'soft': 1.0, 'fail_reason': ''}
    finally:
        shutil.rmtree(work_dir, ignore_errors=True)


# ── Main ────────────────────────────────────────────────────────────────

def main():
    with open(os.path.join(DATA_ROOT, 'dataset.json'), encoding='utf-8') as f:
        items = json.load(f)
    print(f'Items: {len(items)}')

    # Test with initial skill
    skill = "Solve spreadsheet tasks using Python and openpyxl. Write clean, correct code."
    n = 400
    results = []
    t0 = time.time()
    with ThreadPoolExecutor(max_workers=8) as ex:
        futs = {ex.submit(rollout_one, items[i], skill): i for i in range(n)}
        for fut in futs:
            r = fut.result()
            results.append(r)
            i = futs[fut]
            print(f"  [{items[i]['id']}] hard={r['hard']} soft={r['soft']:.3f} {r.get('fail_reason','')[:80]}")

    hard = np.mean([r['hard'] for r in results])
    soft = np.mean([r['soft'] for r in results])
    print(f'\n{n} items: hard={hard:.3f} soft={soft:.3f} ({time.time()-t0:.0f}s)')
    return hard, soft


if __name__ == '__main__':
    main()
